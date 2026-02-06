import openpyxl
import os

filepath = os.path.join(
    'C:/Users/tt/Desktop/Appel d' + chr(39) + 'offres',
    'Construction d' + chr(39) + 'un P' + chr(244) + 'le VNF ' + chr(224) + ' Gambsheim',
    'A mettre en ligne av 19062023',
    'Lot13_Destribois-DPGF.xlsx'
)

print(f'File exists: {os.path.exists(filepath)}')

wb = openpyxl.load_workbook(filepath, data_only=True)
print(f'Sheet names: {wb.sheetnames}')

for name in wb.sheetnames:
    ws = wb[name]
    print('=' * 100)
    print(f'SHEET: {name} (Rows: {ws.max_row}, Cols: {ws.max_column})')
    print('=' * 100)
    if ws.merged_cells.ranges:
        print(f'Merged cells: {[str(m) for m in ws.merged_cells.ranges]}')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=min(ws.max_column, 15), values_only=False):
        values = []
        for cell in row:
            v = cell.value
            if v is not None:
                values.append(f'{cell.column_letter}{cell.row}={repr(v)}')
        if values:
            print(f'Row {row[0].row:3d}: {chr(124).join(values)}')
    print()
